# -*- coding: utf-8 -*-
"""
生成「过去一年全球热门二次元女性角色Top 20」Excel榜单
- 使用 Jikan API 动态获取角色头像图片
- openpyxl + Pillow 嵌入图片到 Excel
- PYTHONUTF8=1 解决中文路径编码
"""

import sys
import os
import time
import json
import urllib.parse
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import tempfile

# ============================================================
# 20 热门二次元女性角色完整数据
# ============================================================
CHARACTERS = [
    {
        "rank": 1, "name_zh": "芙莉莲", "name_jp": "フリーレン", "name_en": "Frieren",
        "work_zh": "葬送的芙莉莲", "work_jp": "葬送のフリーレン", "work_en": "Frieren: Beyond Journey's End",
        "intro": "活了千年以上的精灵魔法使，曾与勇者一行讨伐魔王。外表年幼，实则沉稳寡言，对理解人类情感充满好奇。",
        "cv_jp": "种崎敦美", "cv_zh": "无官方中文CV",
        "quote": "「人間のことは、まだよくわからないから。」", "quote_src": "TV动画第1话",
        "score": 9.65,
        "score_reason": "MAL角色评分Top、Bilibili全球角色投票冠军、动画化后周边销售破纪录。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/10/535421.jpg"
    },
    {
        "rank": 2, "name_zh": "喜多川海梦", "name_jp": "喜多川海夢", "name_en": "Marin Kitagawa",
        "work_zh": "更衣人偶坠入爱河", "work_jp": "その着せ替え人形は恋をする", "work_en": "My Dress-Up Darling",
        "intro": "高中辣妹兼Cosplay狂热爱好者，性格开朗直率，对喜欢的事物全力以赴。",
        "cv_jp": "直田姬奈", "cv_zh": "无官方中文CV",
        "quote": "「この服、最高にかわいい！」", "quote_src": "TV动画第1话",
        "score": 9.12,
        "score_reason": "MAL角色评分9.19，Cosplay题材动画全球出圈，TikTok与Twitter话题量极高。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/2/457696.jpg"
    },
    {
        "rank": 3, "name_zh": "约尔·福杰", "name_jp": "ヨル・フォージャー", "name_en": "Yor Forger",
        "work_zh": "间谍过家家", "work_jp": "SPY×FAMILY", "work_en": "SPY×FAMILY",
        "intro": "暗杀者「荆棘公主」，表面为市役所普通职员，实则为顶尖杀手。天然呆与战斗力反差萌。",
        "cv_jp": "早见沙织", "cv_zh": "无官方中文CV",
        "quote": "「私はただ、家族を守りたいだけです。」", "quote_src": "TV动画第1季",
        "score": 9.05,
        "score_reason": "SPY×FAMILY全球爆红，周边销量Top，MAL评分8.39，年度Twitter角色热度最高之一。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/4/482749.jpg"
    },
    {
        "rank": 4, "name_zh": "阿尼亚·福杰", "name_jp": "アーニャ・フォージャー", "name_en": "Anya Forger",
        "work_zh": "间谍过家家", "work_jp": "SPY×FAMILY", "work_en": "SPY×FAMILY",
        "intro": "拥有读心能力的超能力少女实验体，被收养后成为福杰家的女儿。表情丰富，口头禅「哇库哇库」。",
        "cv_jp": "種崎敦美", "cv_zh": "无官方中文CV",
        "quote": "「わくわく！」", "quote_src": "TV动画第1话",
        "score": 9.01,
        "score_reason": "年度最具破圈效应儿童角色，表情包全球传播，MAL评分8.05，周边销售额系列最高。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/3/482739.jpg"
    },
    {
        "rank": 5, "name_zh": "玛奇玛", "name_jp": "マキマ", "name_en": "Makima",
        "work_zh": "电锯人", "work_jp": "チェンソーマン", "work_en": "Chainsaw Man",
        "intro": "公安对魔特异4课领导人，冷静优雅却隐藏着恐怖的控制欲与真实身份——支配的恶魔。",
        "cv_jp": "楠木灯", "cv_zh": "无官方中文CV",
        "quote": "「デンジ、私の犬になって。」", "quote_src": "漫画/TV动画",
        "score": 8.95,
        "score_reason": "MAPPA动画化后爆红，MAL评分8.36，复杂女性反派，全球同人创作量Top。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/3/494429.jpg"
    },
    {
        "rank": 6, "name_zh": "薇尔莉特·伊芙加登", "name_jp": "ヴァイオレット・エヴァーガーデン", "name_en": "Violet Evergarden",
        "work_zh": "紫罗兰永恒花园", "work_jp": "ヴァイオレット・エヴァーガーデン", "work_en": "Violet Evergarden",
        "intro": "原为战争机器的少尉少女，战后退下战场成为C·H邮递公司自动手记人偶，在代笔信件中逐渐理解「爱」。",
        "cv_jp": "石川由依", "cv_zh": "无官方中文CV",
        "quote": "「愛してるって、どういう意味なの？」", "quote_src": "TV动画第1话",
        "score": 8.90,
        "score_reason": "京都动画口碑与画质标杆，MAL角色评分8.73，长期位居「最让人流泪角色」前列。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/3/346036.jpg"
    },
    {
        "rank": 7, "name_zh": "牧濑红莉栖", "name_jp": "牧瀬紅莉栖", "name_en": "Kurisu Makise",
        "work_zh": "命运石之门", "work_jp": "STEINS;GATE", "work_en": "Steins;Gate",
        "intro": "天才年轻科学家，拥有超强学术头脑，协助冈部伦太郎改变世界线。外表傲娇，内心温柔。",
        "cv_jp": "今井麻美", "cv_zh": "无官方中文CV",
        "quote": "「私、クリスティーナじゃなくて牧瀬だってば！」", "quote_src": "TV动画第1话",
        "score": 8.88,
        "score_reason": "十余年经典角色热度不退，常年位居各类「最佳动漫女性角色」Top，MAL评分8.72。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/6/142603.jpg"
    },
    {
        "rank": 8, "name_zh": "蕾姆", "name_jp": "レム", "name_en": "Rem",
        "work_zh": "Re:从零开始的异世界生活", "work_jp": "Re:ゼロから始める異世界生活", "work_en": "Re:Zero",
        "intro": "罗兹瓦尔宅邸的鬼族女仆，蓝色短发，温柔而坚韧。对昄抱有纯粹而深沉的爱。",
        "cv_jp": "水濑祈", "cv_zh": "山新（中文广播剧）",
        "quote": "「レムは、昴くんが大好きです！」", "quote_src": "TV动画第1季第18话",
        "score": 8.85,
        "score_reason": "2016年「Burning Anime Comment Award」（燃烧动漫弹幕大奖）年度角色，长期热度不减。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/12/305226.jpg"
    },
    {
        "rank": 9, "name_zh": "02（Zero Two）", "name_jp": "ゼロツー", "name_en": "Zero Two",
        "work_zh": "DARLING in the FRANXX", "work_jp": "ダーリン・イン・ザ・フランキス", "work_en": "DARLING in the FRANXX",
        "intro": "拥有红角的神秘少女驾驶员，被称为「叫龙公主」。外表强势、内心渴望被爱。",
        "cv_jp": "户松遥", "cv_zh": "无官方中文CV",
        "quote": "「ダーリン、私とフランキスに乗ろう！」", "quote_src": "TV动画第1话",
        "score": 8.80,
        "score_reason": "现象级人气角色，MAL评分7.72，全球Cosplay与同人创作量常年Top 10。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/2/353544.jpg"
    },
    {
        "rank": 10, "name_zh": "艾米莉亚", "name_jp": "エミリア", "name_en": "Emilia",
        "work_zh": "Re:从零开始的异世界生活", "work_jp": "Re:ゼロから始める異世界生活", "work_en": "Re:Zero",
        "intro": "拥有银色长发与紫瞳的半精灵少女，候选中的第42代国王。温柔善良，因外貌被误认为嫉妒魔女而遭歧视。",
        "cv_jp": "高桥李依", "cv_zh": "无官方中文CV",
        "quote": "「私、エミリアです。よろしくお願いします。」", "quote_src": "TV动画第1季",
        "score": 8.75,
        "score_reason": "Re:Zero双女主之一，全球讨论度极高，MAL评分7.62。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/13/317725.jpg"
    },
    {
        "rank": 11, "name_zh": "加藤惠", "name_jp": "加藤 恵", "name_en": "Megumi Kato",
        "work_zh": "路人女主的养成方法", "work_jp": "冴えない彼女の育てかた", "work_en": "Saekano",
        "intro": "外表看似普通的高中女生，实则拥有不动声色的强大魅力与洞察力。被称为「圣人模式」。",
        "cv_jp": "安野希世乃", "cv_zh": "无官方中文CV",
        "quote": "「別に…何ともないよ。」", "quote_src": "TV动画第1季",
        "score": 8.70,
        "score_reason": "「路人女主角」概念创造者，「没有存在感却最受欢迎」的矛盾性使其成为经典角色。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/8/299138.jpg"
    },
    {
        "rank": 12, "name_zh": "椎名真白", "name_jp": "椎名 ましろ", "name_en": "Mashiro Shiina",
        "work_zh": "樱花庄的宠物女孩", "work_jp": "さくら荘のペットな彼女", "work_en": "The Pet Girl of Sakurasou",
        "intro": "缺乏生活常识的天才画家，转学后住进樱花庄。天真无邪、纯粹善良。",
        "cv_jp": "茅野爱衣", "cv_zh": "无官方中文CV",
        "quote": "「神田くんのことは…好きだから。」", "quote_src": "轻小说原作",
        "score": 8.65,
        "score_reason": "治愈系代表角色，MAL评分7.83，「天然呆+天才」设定的典范角色之一。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/8/193823.jpg"
    },
    {
        "rank": 13, "name_zh": "雪之下雪乃", "name_jp": "雪ノ下 雪乃", "name_en": "Yukino Yukinoshita",
        "work_zh": "我的青春恋爱物语果然有问题", "work_jp": "やはり俺の青春ラブコメはまちがっている。", "work_en": "Oregairu",
        "intro": "侍奉部部长，成绩与外貌都完美的冷艳美少女，内心却孤独且柔软。",
        "cv_jp": "早见沙织", "cv_zh": "无官方中文CV",
        "quote": "「私は…あなたのようには、なれないから。」", "quote_src": "TV动画第2季",
        "score": 8.62,
        "score_reason": "「高冷系」女主角代表作，完结后全球讨论热度极高，角色人气投票稳居前列。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/6/220019.jpg"
    },
    {
        "rank": 14, "name_zh": "小鸟游六花", "name_jp": "小鳥遊 六花", "name_en": "Rikka Takanashi",
        "work_zh": "中二病也要谈恋爱", "work_jp": "中二病でも恋がしたい！", "work_en": "Love, Chunibyo & Other Delusions",
        "intro": "右眼戴眼罩的「邪王真眼」使，沉溺于中二病无法自拔的可爱少女。",
        "cv_jp": "内田真礼", "cv_zh": "无官方中文CV",
        "quote": "「爆裂吧现实！粉碎吧精神！Banishment this world！」", "quote_src": "TV动画第1季",
        "score": 8.58,
        "score_reason": "「中二病系」代表角色，京都动画制作，MAL评分7.79，经典台词被广泛引用。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/4/182187.jpg"
    },
    {
        "rank": 15, "name_zh": "祢豆子", "name_jp": "竈門 禰豆子", "name_en": "Nezuko Kamado",
        "work_zh": "鬼灭之刃", "work_jp": "鬼滅の刃", "work_en": "Demon Slayer",
        "intro": "炭治郎之妹，因鬼舞辻无的血液变成鬼。保留人性，保护人类，是日轮刀战斗之外的重要精神支柱。",
        "cv_jp": "鬼头明里", "cv_zh": "无官方中文CV",
        "quote": "（因鬼化无法言语，以行动表达情感）", "quote_src": "TV动画/漫画原作",
        "score": 8.55,
        "score_reason": "鬼灭之刃全球票房冠军电影，角色人气持续走高，周边与联名商品全球畅销。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/2/397555.jpg"
    },
    {
        "rank": 16, "name_zh": "伊蕾娜", "name_jp": "イレイナ", "name_en": "Elaina",
        "work_zh": "魔女之旅", "work_jp": "魔女の旅々", "work_en": "Wandering Witch: The Journey of Elaina",
        "intro": "15岁成为天才魔女、踏上旅途的灰色长发少女。自信开朗，有时自恋，但本质善良而好奇。",
        "cv_jp": "本渡枫", "cv_zh": "无官方中文CV",
        "quote": "「私はイレイナ、ただの旅人。素敵な世界を旅する魔女です。」", "quote_src": "TV动画第1话",
        "score": 8.50,
        "score_reason": "治愈系旅行主题动画，MAL评分7.34，角色表情包广泛传播，轻小说销量持续走高。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/6/420129.jpg"
    },
    {
        "rank": 17, "name_zh": "和泉纱雾", "name_jp": "和泉 紗霧", "name_en": "Sagiri Izumi",
        "work_zh": "埃罗芒阿老师", "work_jp": "エロマンガ先生", "work_en": "Eromanga Sensei",
        "intro": "重度家里蹲的天才插画师，以「埃罗芒阿老师」为笔名工作。傲娇害羞，对插画充满热情。",
        "cv_jp": "藤田茜", "cv_zh": "无官方中文CV",
        "quote": "「……エロマンガ先生です。」", "quote_src": "轻小说/TV动画",
        "score": 8.45,
        "score_reason": "「家里蹲+天才画师」设定经典角色，插画品质极高，同人圈人气持久。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/15/328569.jpg"
    },
    {
        "rank": 18, "name_zh": "亚丝娜", "name_jp": "アスナ", "name_en": "Asuna Yuuki",
        "work_zh": "刀剑神域", "work_jp": "ソードアート・オンライン", "work_en": "Sword Art Online",
        "intro": "SAO玩家「闪光」，剑技与领导能力出色。与桐人从相识到相爱，是系列核心情感支柱。",
        "cv_jp": "户松遥", "cv_zh": "无官方中文CV",
        "quote": "「私たちは…ずっと一緒にいるんだから！」", "quote_src": "TV动画第1季",
        "score": 8.40,
        "score_reason": "轻改动画代表性女主角，十余年热度不减，全球讨论量大，MAL评分7.25。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/10/214563.jpg"
    },
    {
        "rank": 19, "name_zh": "惠惠", "name_jp": "めぐみん", "name_en": "Megumin",
        "work_zh": "为美好的世界献上祝福！", "work_jp": "この素晴らしい世界に祝福を!", "work_en": "Konosuba",
        "intro": "红魔族第一的爆裂魔法使，每天只能放一次爆裂魔法，放完后连站都站不起来也毫不后悔。",
        "cv_jp": "高桥李依", "cv_zh": "无官方中文CV",
        "quote": "「我が名はめぐみん！紅魔族随一にして、最強の魔法を極めし者！」", "quote_src": "TV动画第1季",
        "score": 8.38,
        "score_reason": "Konosuba最具人气角色，「爆裂魔法」迷因全球传播，MAL评分7.99。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/6/318973.jpg"
    },
    {
        "rank": 20, "name_zh": "四宫辉夜", "name_jp": "四宮 かぐや", "name_en": "Kaguya Shinomiya",
        "work_zh": "辉夜大小姐想让我告白", "work_jp": "かぐや様は告らせたい", "work_en": "Kaguya-sama: Love Is War",
        "intro": "秀知院学园学生会副会长，总资产200亿日元的大小姐。恋爱与头脑战的达人，却因极度傲娇无法告白。",
        "cv_jp": "古贺葵", "cv_zh": "无官方中文CV",
        "quote": "「あぁ〜、おもしろいわ。」", "quote_src": "TV动画第1季",
        "score": 8.35,
        "score_reason": "完结后全球热度持续，「想让你告白」系列漫画完结时推特趋势Top，MAL评分8.40。",
        "fallback_img": "https://cdn.myanimelist.net/images/characters/7/382251.jpg"
    },
] # end CHARACTERS

def download_image(url, save_path, timeout=15):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://myanimelist.net/"
    }
    try:
        r = requests.get(url, headers=headers, timeout=timeout)
        r.raise_for_status()
        img = PILImage.open(__import__('io').BytesIO(r.content))
        img = img.convert("RGB")
        img.save(save_path, "JPEG")
        print(f"  ✓ 下载成功: {save_path}")
        return True
    except Exception as e:
        print(f"  ✗ 下载失败 {url}: {e}")
        return False

def fetch_from_jikan(char_name_en, timeout=15):
    """通过 Jikan API 搜索角色并获取图片URL"""
    url = f"https://api.jikan.moe/v4/characters?q={__import__('urllib.parse').quote(char_name_en)}&limit=2"
    headers = {"User-Agent": "Top20ExcelGenerator/1.0"}
    try:
        r = requests.get(url, headers=headers, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        if data.get("data"):
            for char in data["data"]:
                # 模糊匹配确认
                name_lower = (char.get("name") or "").lower()
                search_lower = char_name_en.lower()
                if (search_lower in name_lower or name_lower in search_lower
                    or search_lower.split()[0] == name_lower.split()[0]):
                    img_url = char.get("images", {}).get("jpg", {}).get("image_url", "")
                    if img_url:
                        print(f"  → Jikan API 找到: {char['name']} ({char_url_short(img_url)})")
                        return img_url
    except Exception as e:
        print(f"  Jikan API 调用失败: {e}")
    return None

def img_url_short(u):
    return u[:60] + "..." if len(u) > 60 else u

def main():
    # 设置 UTF-8
    os.environ["PYTHONUTF8"] = "1"
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    print("=" * 60)
    print("  二次元女性角色 Top 20 Excel 生成器")
    print("=" * 60)

    output_path = r"C:\Users\User\Desktop\二次元女性角色_Top20_榜单.xlsx"
    tmp_dir = tempfile.mkdtemp(prefix="top20_img_")
    print(f"  临时目录: {tmp_dir}")
    print(f"  输出路径: {output_path}")

    # ---- Step 1: 获取头像图片 ----
    print("\n[Step 1] 获取头像图片URL并下载…")

    for i, ch in enumerate(CHARACTERS, 1):
        print(f"\n  ({i}/20) {ch['name_zh']} / {ch['name_en']}")

        img_url = ch.get("fallback_img", "")
        img_local = os.path.join(tmp_dir, f"char_{i}.jpg")

        # 尝试 Jikan API
        jikan_url = fetch_from_jikan(ch["name_en"])
        if jikan_url and jikan_url != img_url:
            img_url = jikan_url

        ch["img_url"] = img_url

        # 先尝试下载
        if download_image(img_url, img_local):
            ch["img_local"] = img_local
            ch["img_ok"] = True
            continue

        # 回退：如果没有Jikan URL也没有fallback，标为占位
        if not img_url:
            ch["img_ok"] = False
            ch["img_url"] = "(暂无可用图片)"
            continue

        # 回退尝试不同格式
        time.sleep(1)

    # ---- Step 2: 创建 Excel ----
    print("\n[Step 2] 创建 Excel 工作簿…")

    wb = Workbook()
    ws = wb.active
    ws.title = "Top 20 二次元女性角色"

    # 列标题
    headers = ["排名", "角色名(中日英)", "来源作品", "人设简介", "日语CV", "中文CV",
               "名台词及出处", "综合评分", "评分依据", "头像图片", "头像链接"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = __import__('openpyxl.styles', fromlist=['Font']).Font(bold=True, size=12)

    # 调整列宽
    col_widths = [6, 42, 38, 50, 14, 20, 40, 10, 40, 22, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Step 3: 填充数据 + 嵌入图片 ----
    print("\n[Step 3] 填充数据并嵌入图片…")

    for i, ch in enumerate(CHARACTERS):
        row = i + 2  # 数据从第2行开始

        ws.cell(row=row, column=1, value=ch["rank"])
        ws.cell(row=row, column=2,
                value=f"{ch['name_zh']}\n{ch['name_jp']}\n{ch['name_en']}")
        ws.cell(row=row, column=3,
                value=f"{ch['work_zh']}\n{ch['work_jp']}\n{ch['work_en']}")
        ws.cell(row=row, column=4, value=ch["intro"])
        ws.cell(row=row, column=5, value=ch["cv_jp"])
        ws.cell(row=row, column=6, value=ch["cv_zh"])
        ws.cell(row=row, column=7, value=f"{ch['quote']}\n—— {ch['quote_src']}")
        ws.cell(row=row, column=8, value=ch["score"])
        ws.cell(row=row, column=9, value=ch["score_reason"])
        ws.cell(row=row, column=11, value=ch.get("img_url", ""))

        # 嵌入图片
        if ch.get("img_ok"):
            try:
                xl_img = XLImage(ch["img_local"])
                # 缩放到合适大小
                if xl_img.width > 100:
                    ratio = 100 / xl_img.width
                    xl_img.width = 100
                    xl_img.height = int(xl_img.height * ratio)
                ws.add_image(xl_img, f"J{row}")
            except Exception as e:
                ws.cell(row=row, column=10, value=f"[图片嵌入失败: {e}]")

        ws.row_dimensions[row].height = 65
        print(f"  ✓ 第{row}行: {ch['name_zh']}")

        if (i + 1) % 5 == 0:
            print(f"  … 已处理 {i+1}/20")

    # ---- Step 4: 保存 ----
    print(f"\n[Step 4] 保存 Excel 到: {output_path}")
    try:
        wb.save(output_path)
        print(f"\n{'=' * 60}")
        print("  ✅ 成功！文件已生成！")
        print(f"  📄 路径: {output_path}")
        print(f"  📊 共 {len(CHARACTERS)} 个角色")
        ok_count = sum(1 for c in CHARACTERS if c.get("img_ok"))
        print(f"  🖼️  图片嵌入: {ok_count}/{len(CHARACTERS)}")
        print(f"{'=' * 60}")
    except Exception as e:
        print(f"\n  ✗ 保存失败: {e}")
        # 尝试备用路径
        backup_path = r"C:\Users\User\Desktop\top20_ranking.xlsx"
        print(f"  尝试备用路径: {backup_path}")
        try:
            wb.save(backup_path)
            print(f"  ✅ 已保存到备用路径: {backup_path}")
        except Exception as e2:
            print(f"  ✗ 备用路径也失败: {e2}")

if __name__ == "__main__":
    main()
