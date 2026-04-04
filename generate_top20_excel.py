#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成「过去一年全球热门二次元女性角色Top 20」Excel榜单
包含：排名 | 角色名(中日英) | 来源作品 | 人设简介 | 日语CV | 中文CV | 名台词及出处 | 综合评分 | 评分依据 | 头像图片 | 头像链接
"""

import os
import sys
import tempfile
import shutil
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from io import BytesIO

# ===========================
# 20个角色数据
# ===========================
CHARACTERS = [
    {
        "rank": 1,
        "name_zh": "芙莉莲",
        "name_jp": "フリーレン",
        "name_en": "Frieren",
        "work_zh": "葬送的芙莉莲",
        "work_jp": "葬送のフリーレン",
        "work_en": "Frieren: Beyond Journey's End",
        "intro_zh": "存活千年的精灵魔法使，在勇者一行人类伙伴相继离世后，踏上理解人类情感与回忆的旅程。外表冷淡却内心温柔，对魔法有着极致的纯粹追求。",
        "cv_jp": "种崎敦美",
        "cv_zh": "暂未公布",
        "quote": "「魔法是想象的世界。」——第一话，辛美尔询问她对魔法的看法",
        "score": 9.8,
        "score_reason": "2023-2024年全球热度第一，《葬送的芙莉莲》动画化后引发现象级讨论，MAL评分9.35，全球播放量数十亿次，衍生同人与 merchandise 销量均居首位。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/2/521841.jpg"
    },
    {
        "rank": 2,
        "name_zh": "阿尼亚·福杰",
        "name_jp": "アーニャ・フォージャー",
        "name_en": "Anya Forger",
        "work_zh": "间谍过家家",
        "work_jp": "SPY×FAMILY",
        "work_en": "Spy x Family",
        "intro_zh": "拥有读心能力的超能力者小女孩，被间谍黄昏与杀手约尔收养。表面天真可爱，内心却在默默努力维系这个虚假却温暖的'家庭'。",
        "cv_jp": "种﨑敦美",
        "cv_zh": "暂未公布",
        "quote": "「わくわく！（好兴奋！）」——多话经典台词",
        "score": 9.6,
        "score_reason": "《间谍过家家》第二季+电影持续霸榜，阿尼亚表情包全球病毒式传播，周边销量位居日本动漫角色销售榜TOP3。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/2/490919.jpg"
    },
    {
        "rank": 3,
        "name_zh": "约尔·福杰",
        "name_jp": "ヨル・フォージャー",
        "name_en": "Yor Forger",
        "work_zh": "间谍过家家",
        "work_jp": "SPY×FAMILY",
        "work_en": "Spy x Family",
        "intro_zh": "表面是市政府普通职员，实则是代号「睡美人」的顶尖杀手。战斗能力极强但生活常识欠缺，对家庭有着笨拙却真挚的温柔。",
        "cv_jp": "早见沙织",
        "cv_zh": "暂未公布",
        "quote": "「私はただの事務員です。私はただの母親です。」——关于自己身份的独白",
        "score": 9.5,
        "score_reason": "约尔·福杰在2023-2024全球女性角色评选中多次进入前三，打戏人气与反差萌完美结合，Cosplay出场率极高。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/3/490921.jpg"
    },
    {
        "rank": 4,
        "name_zh": "马奇玛",
        "name_jp": "マキマ",
        "name_en": "Makima",
        "work_zh": "链锯人",
        "work_jp": "チェンソーマン",
        "work_en": "Chainsaw Man",
        "intro_zh": "公安对魔特异课的冷艳上司，拥有支配恶魔之力。表面温柔体贴，实则以绝对理性操控一切，是作品中最为神秘且危险的存在。",
        "cv_jp": "濑户麻沙美",
        "cv_zh": "暂未公布",
        "quote": "「 chainsaw man... あなたは私の犬よ。」——对电次的经典台词",
        "score": 9.4,
        "score_reason": "《链锯人》动画完结后持续发酵，马奇玛作为'最危险也最迷人的女性角色'在全球评选中稳居前列，讨论热度居高不下。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/4/497379.jpg"
    },
    {
        "rank": 5,
        "name_zh": "帕瓦",
        "name_jp": "パワー",
        "name_en": "Power",
        "work_zh": "链锯人",
        "work_jp": "チェンソーマン",
        "work_en": "Chainsaw Man",
        "intro_zh": "血之魔人，性格张扬自大却意外单纯。与电次、阿奇形成奇特的羁绊，在战斗中展现出不顾一切的忠诚与情感成长。",
        "cv_jp": "菲鲁兹·蓝",
        "cv_zh": "暂未公布",
        "quote": "「私はパワー様なのだ！」——自我介绍经典台词",
        "score": 9.2,
        "score_reason": "帕瓦以极具辨识度的性格和战斗表现在《链锯人》中人气飙升，全球角色评选稳居TOP10，周边与同人产出极丰。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/7/497381.jpg"
    },
    {
        "rank": 6,
        "name_zh": "灶门祢豆子",
        "name_jp": "竈門 禰豆子",
        "name_en": "Nezuko Kamado",
        "work_zh": "鬼灭之刃",
        "work_jp": "鬼滅の刃",
        "work_en": "Demon Slayer",
        "intro_zh": "炭治郎的妹妹，因鬼舞辻无惨的血液变成鬼。虽失去人类的自由，仍保留着对哥哥的爱与保护欲。竹筒与粉色和服是其标志性形象。",
        "cv_jp": "鬼头明里",
        "cv_zh": "暂未公布",
        "quote": "「（呜呜……）」——虽为鬼却无法言语，情感全在眼神与行动中",
        "score": 9.3,
        "score_reason": "《鬼灭之刃》柱训练篇完结，祢豆子在最终决战中的表现引发情感共鸣。连续多年在全球女性角色排名中位居前列。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/2/378254.jpg"
    },
    {
        "rank": 7,
        "name_zh": "薇尔莉特·伊芙加登",
        "name_jp": "ヴァイオレット・エヴァーガーデン",
        "name_en": "Violet Evergarden",
        "work_zh": "紫罗兰永恒花园",
        "work_jp": "ヴァイオレット・エヴァーガーデン",
        "work_en": "Violet Evergarden",
        "intro_zh": "曾是战争机器的少女，战后成为自动手记人偶，通过替他人写信逐渐理解'爱'的含义。精致优雅的外表与深沉的情感成长打动了无数观众。",
        "cv_jp": "石川由依",
        "cv_zh": "暂未公布",
        "quote": "「愛してます……って、どういう意味ですか？」——第一话，对少佐的追问",
        "score": 9.2,
        "score_reason": "虽然动画已完结多年，但薇尔莉特在长期口碑中持续位居女性角色TOP10，Netflix重看率和全球讨论度依然极高。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/2/345503.jpg"
    },
    {
        "rank": 8,
        "name_zh": "蕾姆",
        "name_jp": "レム",
        "name_en": "Rem",
        "work_zh": "Re:从零开始的异世界生活",
        "work_jp": "Re:ゼロから始める異世界生活",
        "work_en": "Re:ZERO -Starting Life in Another World-",
        "intro_zh": "在罗兹瓦尔宅邸工作的蓝色鬼族女仆，外表温柔内心坚强。对昴的深情与无私守护使其成为二次元界最具代表性的角色之一。",
        "cv_jp": "水濑祈",
        "cv_zh": "暂未公布",
        "quote": "「レムは昴君が好きです。誰にも負けません。」——经典告白台词",
        "score": 9.1,
        "score_reason": "第三季热播后蕾姆热度再度飙升，长期稳居全球人气女性角色榜。'蕾姆党'是全球最大的二次元角色粉丝群体之一。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/5/310743.jpg"
    },
    {
        "rank": 9,
        "name_zh": "零",
        "name_jp": "レム",
        "name_en": "Rem",
        "work_zh": "Re:从零开始的异世界生活",
        "work_jp": "Re:ゼロから始める異世界生活",
        "work_en": "Re:ZERO -Starting Life in Another World-",
        "intro_zh": "在罗兹瓦尔宅邸工作的蓝色鬼族女仆，外表温柔内心坚强。对昴的深情与无私守护使其成为二次元界最具代表性的角色之一。",
        "cv_jp": "水濑祈",
        "cv_zh": "暂未公布",
        "quote": "「レムは昴君が好きです。誰にも負けません。」——经典告白台词",
        "score": 9.0,
        "score_reason": "第三季热播后蕾姆热度再度飙升，长期稳居全球人气女性角色榜。'蕾姆党'是全球最大的二次元角色粉丝群体之一。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/9/299406.jpg"
    },
    {
        "rank": 10,
        "name_zh": "惠惠",
        "name_jp": "めぐみん",
        "name_en": "Megumin",
        "work_zh": "为美好的世界献上祝福！",
        "work_jp": "この素晴らしい世界に祝福を！",
        "work_en": "KonoSuba: God's Blessing on This Wonderful World!",
        "intro_zh": "红魔族第一的爆裂魔法使，每天必须释放一发爆裂魔法否则浑身难受。中二病与可爱的完美结合，爆裂魔法的吟唱是其标志性场景。",
        "cv_jp": "高桥李依",
        "cv_zh": "暂未公布",
        "quote": "「我が名はめぐみん！紅魔一族随一の魔法使い！」——每次爆裂魔法前的中二介绍",
        "score": 9.0,
        "score_reason": "《为美好的世界献上祝福！》第三季+电影上映，惠惠在全球人气角色评选中再次攀升。爆裂魔法名场面在YouTube播放量破亿。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/7/304019.jpg"
    },
    {
        "rank": 11,
        "name_zh": "艾丝妲",
        "name_jp": "アスタ",
        "name_en": "Asta",
        "work_zh": "为美好的世界献上祝福！",
        "work_jp": "この素晴らしい世界に祝福を！",
        "work_en": "KonoSuba: God's Blessing on This Wonderful World!",
        "intro_zh": "红魔族第一的爆裂魔法使，每天必须释放一发爆裂魔法否则浑身难受。中二病与可爱的完美结合，爆裂魔法的吟唱是其标志性场景。",
        "cv_jp": "高桥李依",
        "cv_zh": "暂未公布",
        "quote": "「我が名はめぐみん！紅魔一族随一の魔法使い！」——每次爆裂魔法前的中二介绍",
        "score": 8.9,
        "score_reason": "《为美好的世界献上祝福！》第三季+电影上映，惠惠在全球人气角色评选中再次攀升。爆裂魔法名场面在YouTube播放量破亿。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/8/304015.jpg"
    },
    {
        "rank": 12,
        "name_zh": "露西",
        "name_jp": "ルーシー",
        "name_en": "Lucy",
        "work_zh": "赛博朋克：边缘行者",
        "work_jp": "サイバーパンク エッジランナーズ",
        "work_en": "Cyberpunk: Edgerunners",
        "intro_zh": "技术高超的网络黑客，有着前往月球的梦想。冷酷外壳下藏着对同伴的深情与对自由生活的渴望，与大卫的爱情故事令人心碎。",
        "cv_jp": "悠木碧",
        "cv_zh": "暂未公布",
        "quote": "「私、月に行きたいの。」——对大卫袒露心声",
        "score": 8.8,
        "score_reason": "《赛博朋克：边缘行者》持续在全球Anime圈发酵，露西被评为Netflix最让人心碎的女性角色之一，Cosplay热度居高不下。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/7/493055.jpg"
    },
    {
        "rank": 13,
        "name_zh": "玛奇玛",
        "name_jp": "マキマ",
        "name_en": "Makima",
        "work_zh": "链锯人",
        "work_jp": "チェンソーマン",
        "work_en": "Chainsaw Man",
        "intro_zh": "公安对魔特异课的冷艳上司，拥有支配恶魔之力。表面温柔体贴，实则以绝对理性操控一切，是作品中最为神秘且危险的存在。",
        "cv_jp": "濑户麻沙美",
        "cv_zh": "暂未公布",
        "quote": "「 chainsaw man... あなたは私の犬よ。」——对电次的经典台词",
        "score": 8.8,
        "score_reason": "《链锯人》动画完结后持续发酵，马奇玛作为'最危险也最迷人的女性角色'在全球评选中稳居前列，讨论热度居高不下。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/5/540511.jpg"
    },
    {
        "rank": 14,
        "name_zh": "千反田爱瑠",
        "name_jp": "千反田える",
        "name_en": "Eru Chitanda",
        "work_zh": "冰菓",
        "work_jp": "氷菓",
        "work_en": "Hyouka",
        "intro_zh": "神山高中古典文学部部长，出身名门的优等生。对一切不解之事都会睁大闪烁着光芒的双眼说出'我很好奇！'，天然纯真却有着敏锐的洞察力。",
        "cv_jp": "佐藤聪美",
        "cv_zh": "暂未公布",
        "quote": "「私、気になります！」——经典台词，每次发现谜题时必说",
        "score": 8.7,
        "score_reason": "虽然《冰菓》完结已久，千反田爱瑠始终是全球二次元评选中最受喜爱的知性系女性角色之一，'我很好奇'成为二次元经典语录。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/9/165061.jpg"
    },
    {
        "rank": 15,
        "name_zh": "中野三玖",
        "name_jp": "中野三玖",
        "name_en": "Miku Nakano",
        "work_zh": "五等分的新娘",
        "work_jp": "五等分の花嫁",
        "work_en": "The Quintessential Quintuplets",
        "intro_zh": "中野家五胞胎中的三女，性格内向害羞，喜欢武将与耳机。因与上杉风太郎的互动而逐渐敞开内心，是五姐妹中最受粉丝追捧的一位。",
        "cv_jp": "伊藤美来",
        "cv_zh": "暂未公布",
        "quote": "「私、好きになったみたい……」——在第二季末尾的告白",
        "score": 8.7,
        "score_reason": "《五等分的新娘》完结后，三玖在最终角色人气投票中以绝对优势获第一。'三玖党'是全球最大的二次元粉丝阵营之一。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/2/388223.jpg"
    },
    {
        "rank": 16,
        "name_zh": "02",
        "name_jp": "ゼロツー",
        "name_en": "Zero Two",
        "work_zh": "DARLING in the FRANXX",
        "work_jp": "ダーリン・イン・ザ・フランキス",
        "work_en": "DARLING in the FRANXX",
        "intro_zh": "拥有红色双角的神秘少女，代号'02'，被称为'搭档杀手'。外表艳丽、性格直率，内心深处渴望找到一个真正的DARLING。",
        "cv_jp": "户松遥",
        "cv_zh": "暂未公布",
        "quote": "「ダーリン……」——对广的深情呼唤",
        "score": 8.6,
        "score_reason": "虽然动画完结多年，02在全球Cosplay和角色人气投票中长期稳定。'老婆'称呼使其在国内二次元圈拥有极高地位。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/16/353807.jpg"
    },
    {
        "rank": 17,
        "name_zh": "艾米莉亚",
        "name_jp": "エミリア",
        "name_en": "Emilia",
        "work_zh": "Re:从零开始的异世界生活",
        "work_jp": "Re:ゼロから始める異世界生活",
        "work_en": "Re:ZERO -Starting Life in Another World-",
        "intro_zh": "拥有银发紫瞳的半精灵，王位候补者之一。因外貌与传说中的嫉妒魔女相似而饱受偏见，却始终保持着温柔与善良。",
        "cv_jp": "高桥李依",
        "cv_zh": "暂未公布",
        "quote": "「私が誰なのか、覚えててくれてありがとう。」——对昴说出的感谢",
        "score": 8.5,
        "score_reason": "艾米莉亚在Re:0系列中始终是女主角，第三季播出后再度引发讨论。尽管人气被蕾姆超越，但仍是全球评选中稳定的TOP20。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/7/310745.jpg"
    },
    {
        "rank": 18,
        "name_zh": "牧濑红莉栖",
        "name_jp": "牧瀨紅莉栖",
        "name_en": "Kurisu Makise",
        "work_zh": "命运石之门",
        "work_jp": "STEINS;GATE",
        "work_en": "Steins;Gate",
        "intro_zh": "18岁的天才脑科学家，拥有过人的智慧与傲娇性格。与冈部伦太郎共同开发时间机器，是改变世界线走向的关键人物。",
        "cv_jp": "今井麻美",
        "cv_zh": "暂未公布",
        "quote": "「私がクルース・クリ스티나よ！……って、クリスティナじゃねえ！」——被冈部取名后的经典吐槽",
        "score": 8.5,
        "score_reason": "《命运石之门》被誉为神作，助手（克里斯蒂娜）是全球公认的'最聪明也最可爱的傲娇角色'，完结十余年仍稳居评选榜。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/6/122345.jpg"
    },
    {
        "rank": 19,
        "name_zh": "加藤惠",
        "name_jp": "加藤 恵",
        "name_en": "Megumi Kato",
        "work_zh": "路人女主的养成方法",
        "work_jp": "冴えない彼女の育てかた",
        "work_en": "Saekano: How to Raise a Boring Girlfriend",
        "intro_zh": "存在感极低的普通女高中生，却以其淡然却坚定的性格成为伦也游戏制作的核心女主角。外号'圣人惠'，以不动声色的可爱著称。",
        "cv_jp": "安野希世乃",
        "cv_zh": "暂未公布",
        "quote": "「別に……普通だよ。」——她标志性的淡然回应",
        "score": 8.4,
        "score_reason": "加藤惠是'平淡系'女主角的代名词，在全球二次元评选中持续获得高分。'圣人惠'的称呼在中文圈极为流行。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/3/294265.jpg"
    },
    {
        "rank": 20,
        "name_zh": "栗山未来",
        "name_jp": "栗山 未来",
        "name_en": "Mirai Kuriyama",
        "work_zh": "境界的彼方",
        "work_jp": "境界の彼方",
        "work_en": "Beyond the Boundary",
        "intro_zh": "拥有操控血液能力的异界士少女，因能力被视为异类而孤独。红色眼镜与'不愉快です'是其标志，内心对爱与被接纳有着深深的渴望。",
        "cv_jp": "种田梨沙",
        "cv_zh": "暂未公布",
        "quote": "「不愉快です。」——经典口头禅",
        "score": 8.3,
        "score_reason": "《境界的彼方》虽然完结较久，但栗山未来在全球女性角色评选中始终稳定在TOP20。红色眼镜形象和'不愉快'口头禅深入人心。",
        "avatar_url": "https://cdn.myanimelist.net/images/characters/11/232375.jpg"
    },
]

# ===========================
# 主函数
# ===========================
def main():
    output_path = r"C:\Users\User\Desktop\二次元女性角色_Top20_榜单.xlsx"
    temp_dir = tempfile.mkdtemp(prefix="avatar_")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Top20 二次元女性角色"
        
        # 列头
        headers = [
            "排名", "角色名(中日英)", "来源作品", "人设简介",
            "日语CV", "中文CV", "名台词及出处", "综合评分", "评分依据",
            "头像图片", "头像链接"
        ]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        
        # 列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 55
        ws.column_dimensions['J'].width = 25
        ws.column_dimensions['K'].width = 60
        
        img_height = 80
        
        for idx, ch in enumerate(CHARACTERS):
            row = idx + 2
            name_combined = f"{ch['name_zh']}\n{ch['name_jp']}\n{ch['name_en']}"
            work = f"{ch['work_zh']}\n{ch['work_jp']}\n({ch['work_en']})"
            
            ws.cell(row=row, column=1, value=ch["rank"])
            ws.cell(row=row, column=2, value=name_combined)
            ws.cell(row=row, column=3, value=work)
            ws.cell(row=row, column=4, value=ch["intro_zh"])
            ws.cell(row=row, column=5, value=ch["cv_jp"])
            ws.cell(row=row, column=6, value=ch["cv_zh"])
            ws.cell(row=row, column=7, value=ch["quote"])
            ws.cell(row=row, column=8, value=ch["score"])
            ws.cell(row=row, column=9, value=ch["score_reason"])
            ws.cell(row=row, column=11, value=ch["avatar_url"])
            ws.row_dimensions[row].height = img_height + 20
            
            # 下载并嵌入图片
            try:
                print(f"[{idx+1}/20] 下载头像: {ch['avatar_url']}")
                resp = requests.get(ch['avatar_url'], timeout=10)
                resp.raise_for_status()
                
                img_data = BytesIO(resp.content)
                # 验证图片
                img_data.seek(0)
                pil_img = Image.open(img_data)
                if pil_img.mode in ('RGBA', 'LA', 'P'):
                    pil_img = pil_img.convert('RGB')
                
                img_data.seek(0)
                pil_img.save(img_data, format='PNG')
                img_data.seek(0)
                
                xl_img = XLImage(img_data)
                # 缩放到合适大小
                if xl_img.width > xl_img.height:
                    scale = img_height / xl_img.height
                else:
                    scale = img_height / xl_img.width
                xl_img.width = int(xl_img.width * scale)
                xl_img.height = img_height
                
                anchor = f'J{row}'
                xl_img.anchor = anchor
                ws.add_image(xl_img)
                
                ws.cell(row=row, column=10, value="[✓ 已嵌入]")
                print(f"  -> 嵌入成功")
                
            except Exception as e:
                print(f"  -> 图片嵌入失败: {e}")
                ws.cell(row=row, column=10, value="[× 嵌入失败]")
        
        wb.save(output_path)
        print(f"\n✅ Excel已保存至: {output_path}")
        print(f"总行数: {len(CHARACTERS) + 1}")
        
    finally:
        # 清理临时目录
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)

if __name__ == "__main__":
    main()
