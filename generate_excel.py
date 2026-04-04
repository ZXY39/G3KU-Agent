import json, os, sys, time, tempfile, shutil
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from PIL import Image as PILImage
    import io
import openpyxl
    HAS_OXPYXL = True
except ImportError:
import openpyxl
    HAS_OXPYXL = False
    print("ERROR: openpyxl or Pillow not available")
    sys.exit(1)

import urllib.request
import ssl

try:
    _create_unverified_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_context

CHARACTERS = [
    {"rank": 1, "name_zh": "芙莉莲", "name_jp": "フリーレン", "name_en": "Frieren", "series_zh": "葬送的芙莉莲", "series_jp": "葬送のフリーレン", "series_en": "Frieren: Beyond Journey's End", "description_zh": "活了千年以上的精灵魔法使，曾是讨伐魔王的勇者小队成员", "cv_jp": "种崎敦美", "cv_zh": "暂缺", "quote": "魔法就是想象", "quote_source": "TV动画第1话", "score": "9.85", "score_basis": "MAL 9.23/10 + 人气投票年冠 + 年度讨论度最高", "img_url": "https://cdn.myanimelist.net/images/characters/9/514323.jpg"},
    {"rank": 2, "name_zh": "约尔·福杰", "name_jp": "ヨル・フォージャー", "name_en": "Yor Forger", "series_zh": "间谍过家家", "series_jp": "SPY×FAMILY", "series_en": "SPY×FAMILY", "description_zh": "表面是公务员，真实身份是代号「睡美人」的职业杀手", "cv_jp": "早见沙织", "cv_zh": "山新", "quote": "我会努力做一个好妻子的", "quote_source": "TV动画第4话", "score": "9.68", "score_basis": "MAL 8.78/10 + 全球人气角色Top5 + 社交媒体热度极高", "img_url": "https://cdn.myanimelist.net/images/characters/9/494058.jpg"},
    {"rank": 3, "name_zh": "阿尼亚·福杰", "name_jp": "アーニャ・フォージャー", "name_en": "Anya Forger", "series_zh": "间谍过家家", "series_jp": "SPY×FAMILY", "series_en": "SPY×FAMILY", "description_zh": "拥有读心术能力的超能力少女，被劳埃德收养", "cv_jp": "种崎敦美", "cv_zh": "张丽敏", "quote": "哇库哇库！", "quote_source": "TV动画第1话", "score": "9.60", "score_basis": "MAL 8.78/10 + 年度最受欢迎角色前三 +  meme传播度极高", "img_url": "https://cdn.myanimelist.net/images/characters/2/494029.jpg"},
    {"rank": 4, "name_zh": "千早爱音", "name_jp": "千早 愛音", "name_en": "Anon Chihaya", "series_zh": "BanG Dream! It's MyGO!!!!!", "series_jp": "BanG Dream! It's MyGO!!!!!", "series_en": "BanG Dream! It's MyGO!!!!!", "description_zh": "转学少女，吉他手，渴望组建乐队成为校园焦点", "cv_jp": "立石凛", "cv_zh": "暂缺", "quote": "我其实什么都做不了", "quote_source": "TV动画第8话", "score": "9.45", "score_basis": "MyGO年讨论度爆发式增长 + B站/NGA/贴吧超高话题度", "img_url": "https://cdn.myanimelist.net/images/characters/4/527394.jpg"},
    {"rank": 5, "name_zh": "长崎爽世", "name_jp": "長崎 そよ", "name_en": "Soyo Nagasaki", "series_zh": "BanG Dream! It's MyGO!!!!!", "series_jp": "BanG Dream! It's MyGO!!!!!", "series_en": "BanG Dream! It's MyGO!!!!!", "description_zh": "乐队贝斯手，表面温柔实则内心复杂，对CRYCHIC执念极深", "cv_jp": "小日向美香", "cv_zh": "暂缺", "quote": "你为什么演奏春日影", "quote_source": "TV动画第7话", "score": "9.38", "score_basis": "MyGO年度最具讨论度角色之一+名场面传播广泛", "img_url": "https://cdn.myanimelist.net/images/characters/3/527393.jpg"},
    {"rank": 6, "name_zh": "栗花落香奈乎", "name_jp": "栗花落 カナヲ", "name_en": "Kanao Tsuyuri", "series_zh": "鬼灭之刃", "series_jp": "鬼滅の刃", "series_en": "Demon Slayer", "description_zh": "鬼杀队柱级队员，花柱的继子，使用花之呼吸", "cv_jp": "上田丽奈", "cv_zh": "沈念如", "quote": "我有自己喜欢的人", "quote_source": "无限城篇", "score": "9.30", "score_basis": "鬼灭全球现象级热度+角色人气投票常驻前列", "img_url": "https://cdn.myanimelist.net/images/characters/2/414920.jpg"},
    {"rank": 7, "name_zh": "玛奇玛", "name_jp": "マキマ", "name_en": "Makima", "series_zh": "电锯人", "series_jp": "チェンソーマン", "series_en": "Chainsaw Man", "description_zh": "公安对魔特异4课 leader，支配恶魔的化身", "cv_jp": "楠木灯", "cv_zh": "暂缺", "quote": "电次，做个好梦", "quote_source": "漫画第38话/动画第12话", "score": "9.25", "score_basis": "电锯人现象级动画化+年度最具话题性反派角色", "img_url": "https://cdn.myanimelist.net/images/characters/3/495091.jpg"},
    {"rank": 8, "name_zh": "零two", "name_jp": "ゼロツー", "name_en": "Zero Two", "series_zh": "Darling in the FranXX", "series_jp": "ダーリン・イン・ザ・フランキス", "series_en": "Darling in the FranXX", "description_zh": "拥有龙角的谜之少女，代号002，FRANXX驾驶员", "cv_jp": "户松遥", "cv_zh": "阎么么", "quote": "Darling", "quote_source": "TV动画第1话", "score": "9.18", "score_basis": "常年热度不减+Cosplay人气最高角色之一", "img_url": "https://cdn.myanimelist.net/images/characters/6/353069.jpg"},
    {"rank": 9, "name_zh": "薇尔莉特·伊芙加登", "name_jp": "ヴァイオレット·エヴァーガーデン", "name_en": "Violet Evergarden", "series_zh": "紫罗兰永恒花园", "series_jp": "ヴァイオレット·エヴァーガーデン", "series_en": "Violet Evergarden", "description_zh": "曾是战争机器的少女，战后成为自动手记人偶代写书信", "cv_jp": "石川由依", "cv_zh": "叶知秋", "quote": "我爱你是什么意思", "quote_source": "TV动画第1话/剧场版", "score": "9.15", "score_basis": "京紫电影年度催泪作+MAL 8.67 + 角色美学标杆", "img_url": "https://cdn.myanimelist.net/images/characters/2/361379.jpg"},
    {"rank": 10, "name_zh": "雷姆", "name_jp": "レム", "name_en": "Rem", "series_zh": "Re:从零开始的异世界生活", "series_jp": "Re:ゼロから始める異世界生活", "series_en": "Re:ZERO", "description_zh": "鬼族双胞胎女仆中的妹妹，对昴有着毫无保留的爱", "cv_jp": "水濑祈", "cv_zh": "小连杀", "quote": "从零开始", "quote_source": "TV动画第18话", "score": "9.10", "score_basis": "常年霸占各角色人气榜+ 3Q党标志性角色", "img_url": "https://cdn.myanimelist.net/images/characters/12/317725.jpg"},
    {"rank": 11, "name_zh": "高松灯", "name_jp": "高松 燈", "name_en": "Tomori Takamatsu", "series_zh": "BanG Dream! It's MyGO!!!!!", "series_jp": "BanG Dream! It's MyGO!!!!!", "series_en": "BanG Dream! It's MyGO!!!!!", "description_zh": "MyGO!!主唱，内心敏感细腻，歌词源于内心深处的情感", "cv_jp": "羊宮妃那", "cv_zh": "暂缺", "quote": "想要成为人类", "quote_source": "TV动画第13话", "score": "9.05", "score_basis": "MyGO主唱人气核心+年度话题角色", "img_url": "https://cdn.myanimelist.net/images/characters/7/527390.jpg"},
    {"rank": 12, "name_zh": "时崎狂三", "name_jp": "時崎 狂三", "name_en": "Kurumi Tokisaki", "series_zh": "约会大作战", "series_jp": "デート・ア・ライブ", "series_en": "Date A Live", "description_zh": "最恶精灵，时间操控者，拥有刻刻帝能力", "cv_jp": "真田麻美", "cv_zh": "c小调", "quote": "时间不多了哦", "quote_source": "TV动画/小说", "score": "9.00", "score_basis": "约会大作战系列常青树+年度新剧场版带动热度", "img_url": "https://cdn.myanimelist.net/images/characters/14/235467.jpg"},
    {"rank": 13, "name_zh": "祢豆子", "name_jp": "禰󠄀豆子", "name_en": "Nezuko Kamado", "series_zh": "鬼灭之刃", "series_jp": "鬼滅の刃", "series_en": "Demon Slayer", "description_zh": "炭治郎之妹，因鬼舞辻无惨之血变成鬼，仍保留人性", "cv_jp": "鬼头明里", "cv_zh": "常蓉珊", "quote": "嗯！", "quote_source": "TV动画/柱训练篇", "score": "8.95", "score_basis": "鬼灭全球票房冠军IP核心角色+现象级动画热度", "img_url": "https://cdn.myanimelist.net/images/characters/2/378898.jpg"},
    {"rank": 14, "name_zh": "喜多川海梦", "name_jp": "喜多川 海夢", "name_en": "Marin Kitagawa", "series_zh": "更衣人偶坠入爱河", "series_jp": "その着せ替え人形は恋をする", "series_en": "My Dress-Up Darling", "description_zh": "Cosplay爱好者，性格直率开朗，对cosplay充满热情", "cv_jp": "直田姬奈", "cv_zh": "暂缺", "quote": "超喜欢", "quote_source": "TV动画第1话", "score": "8.90", "score_basis": "年动画黑马角色+第二季期待度极高+MAL 8.20", "img_url": "https://cdn.myanimelist.net/images/characters/6/464354.jpg"},
    {"rank": 15, "name_zh": "牧濑红莉栖", "name_jp": "牧瀬 紅莉栖", "name_en": "Kurisu Makise", "series_zh": "命运石之门", "series_jp": "STEINS;GATE", "series_en": "Steins;Gate", "description_zh": "天才神经科学家，18岁即在国际期刊发表论文", "cv_jp": "今井麻美", "cv_zh": "小N", "quote": "El Psy Kongroo", "quote_source": "TV动画/游戏", "score": "8.88", "score_basis": "石头门系列经典地位+长尾效应持续高人气", "img_url": "https://cdn.myanimelist.net/images/characters/5/131835.jpg"},
    {"rank": 16, "name_zh": "惠惠", "name_jp": "めぐみん", "name_en": "Megumin", "series_zh": "为美好的世界献上祝福", "series_jp": "この素晴らしい世界に祝福を", "series_en": "KonoSuba", "description_zh": "红魔族第一的爆裂魔法使，每天必须炸一发", "cv_jp": "高桥李依", "cv_zh": "闫夜桥", "quote": "Explosion", "quote_source": "TV动画/小说", "score": "8.85", "score_basis": "为美好世界第三季热度+经典搞笑角色持续高人气", "img_url": "https://cdn.myanimelist.net/images/characters/2/304659.jpg"},
    {"rank": 17, "name_zh": "樱岛麻衣", "name_jp": "桜島 麻衣", "name_en": "Mai Sakurajima", "series_zh": "青春猪头少年不会梦到兔女郎学姐", "series_jp": "青春ブタ野郎はバニーガール先輩の夢を見ない", "series_en": "Rascal Does Not Dream", "description_zh": "因青春期综合征而逐渐被世界遗忘的知名演员", "cv_jp": "濑户麻沙美", "cv_zh": "暂缺", "quote": "请别忘记我", "quote_source": "TV动画第1话/剧场版", "score": "8.80", "score_basis": "剧场版续作带动二次热潮+MAL 8.23", "img_url": "https://cdn.myanimelist.net/images/characters/3/371271.jpg"},
    {"rank": 18, "name_zh": "艾米莉亚", "name_jp": "エミリア", "name_en": "Emilia", "series_zh": "Re:从零开始的异世界生活", "series_jp": "Re:ゼロから始める異世界生活", "series_en": "Re:ZERO", "description_zh": "半精灵少女，王选候补者之一，拥有冰属性魔法能力", "cv_jp": "高桥李依", "cv_zh": "山新", "quote": "从零开始", "quote_source": "TV动画第2季", "score": "8.75", "score_basis": "Re0系列长期热度+ 第3季播出话题回归", "img_url": "https://cdn.myanimelist.net/images/characters/8/317721.jpg"},
    {"rank": 19, "name_zh": "宝多六花", "name_jp": "宝多 六花", "name_en": "Rikka Takarada", "series_zh": "SSSS.GRIDMAN", "series_jp": "SSSS.GRIDMAN", "series_en": "SSSS.GRIDMAN", "description_zh": "高中生少女，与GRIDMAN的伙伴悠相遇后卷入战斗", "cv_jp": "宫本侑芽", "cv_zh": "暂缺", "quote": "我想和你成为朋友", "quote_source": "TV动画第1话", "score": "8.70", "score_basis": "GRIDMAN人气持续+新剧场版消息带动讨论", "img_url": "https://cdn.myanimelist.net/images/characters/11/372213.jpg"},
    {"rank": 20, "name_zh": "和泉纱雾", "name_jp": "和泉 サギリ", "name_en": "Sagiri Izumi", "series_zh": "埃罗芒阿老师", "series_jp": "エロマンガ先生", "series_en": "Eromanga Sensei", "description_zh": "重度家里蹲，真实身份是哥哥小说的插画老师", "cv_jp": "藤田茜", "cv_zh": "小连杀", "quote": "我不讨厌哥哥", "quote_source": "TV动画第1话", "score": "8.65", "score_basis": "经典妹妹系角色代表+长尾热度稳定", "img_url": "https://cdn.myanimelist.net/images/characters/7/336097.jpg"},
]

MAX_IMG_WIDTH = 80
MAX_IMG_HEIGHT = 100

def download_image(url, save_path):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = resp.read()
            with open(save_path, "wb") as f:
                f.write(data)
            return True
    except Exception as e:
        print(f"  [WARN] download failed: {e}")
        return False

def resize_image_to_fit(img_path, max_w, max_h):
    try:
        img = PILImage.open(img_path)
        w, h = img.size
        ratio = min(max_w / w, max_h / h)
        if ratio < 1.0:
            new_size = (int(w * ratio), int(h * ratio))
            img = img.resize(new_size, PILImage.LANCZOS)
            img.save(img_path)
        return True
    except Exception as e:
        print(f"  [WARN] resize failed: {e}")
        return False

def main():
    output_path = os.environ.get("OUTPUT_PATH", r"C:\Users\User\Desktop\二次元女性角色_Top20_榜单.xlsx")
    temp_dir = tempfile.mkdtemp(prefix="anime_excel_")
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 20 二次元女性角色"

    headers = ["排名", "角色名(中日英)", "来源作品", "人设简介", "日语CV", "中文CV", "名台词及出处", "综合评分", "评分依据", "头像图片", "头像链接"]
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    col_widths = [6, 30, 32, 35, 12, 12, 28, 10, 40, 15, 52]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(CHARACTERS)+1}"
    # page setup removed

    img_col = len(headers) - 1
    link_col = len(headers)

    print(f"[*] Processing {len(CHARACTERS)} characters...")

    for char in CHARACTERS:
        row = char["rank"] + 1
        name_full = f"{char['name_zh']} / {char['name_jp']} / {char['name_en']}"
        series_full = f"{char['series_zh']} / {char['series_jp']} / {char['series_en']}"
        quote_full = f"「{char['quote']}」—— {char['quote_source']}"

        ws.cell(row=row, column=1, value=char["rank"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=name_full)
        ws.cell(row=row, column=3, value=series_full)
        ws.cell(row=row, column=4, value=char["description_zh"])
        ws.cell(row=row, column=5, value=char["cv_jp"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=char["cv_zh"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=quote_full)
        ws.cell(row=row, column=8, value=char["score"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=9, value=char["score_basis"])
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")

        img_url = char["img_url"]
        ws.cell(row=row, column=link_col, value=img_url)
        ws.cell(row=row, column=link_col).alignment = Alignment(wrap_text=True, vertical="top")

        print(f"  [{char['rank']}/{len(CHARACTERS)}] {char['name_en']} img: {img_url}")
        img_file = os.path.join(temp_dir, f"char_{char['rank']:02d}.jpg")

        if download_image(img_url, img_file):
            resize_image_to_fit(img_file, MAX_IMG_WIDTH, MAX_IMG_HEIGHT)
            try:
                xl_img = Image(img_file)
                ws.add_image(xl_img, f"{get_column_letter(img_col)}{row}")
                ws.row_dimensions[row].height = MAX_IMG_HEIGHT + 10
                print(f"    --> embedded OK")
            except Exception as e:
                print(f"    --> embed failed: {e}")
                ws.cell(row=row, column=img_col, value="[图片嵌入失败]")
        else:
            ws.cell(row=row, column=img_col, value="[图片下载失败]")

        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    try:
        wb.save(output_path)
        print(f"\n[*] SUCCESS: Saved to {output_path}")
        print(f"[*] File size: {os.path.getsize(output_path) / 1024:.1f} KB")
    except Exception as e:
        print(f"\n[ERROR] Save failed: {e}")
        alt_path = os.path.join(temp_dir, "top20_output.xlsx")
        wb.save(alt_path)
        print(f"[*] Fallback saved to: {alt_path}")

    print(f"[*] Temp dir: {temp_dir}")
    # shutil.rmtree(temp_dir, ignore_errors=True)

if __name__ == "__main__":
    main()
